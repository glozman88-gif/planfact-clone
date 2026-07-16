"""Экспорт отчётов и операций в Excel (.xlsx) через openpyxl.

Принимает готовые словари отчётов (из app.services.reports) и собирает книгу.
Денежные значения приходят строками (Decimal) — в ячейки кладём числа (float),
форматирование выполняет Excel по числовому формату.
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

BRAND = "16B1BF"
HEADER_FILL = PatternFill("solid", fgColor=BRAND)
HEADER_FONT = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
SUBTLE = Font(color="64748B")
RIGHT = Alignment(horizontal="right")
MONEY = "#,##0.00;[Red]-#,##0.00"


def _num(v) -> float:
    """Строку/Decimal/None → float для числовой ячейки Excel."""
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _money_cell(ws: Worksheet, row: int, col: int, value, *, bold: bool = False):
    c = ws.cell(row=row, column=col, value=_num(value))
    c.number_format = MONEY
    c.alignment = RIGHT
    if bold:
        c.font = BOLD
    return c


def _header_row(ws: Worksheet, row: int, labels: list[str]):
    for col, label in enumerate(labels, start=1):
        c = ws.cell(row=row, column=col, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        if col > 1:
            c.alignment = RIGHT


def _autosize(ws: Worksheet, widths: dict[int, int]):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _save(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ───────────────────────────── ОПиУ ─────────────────────────────

def _flatten_categories(cats: list[dict], depth: int = 0):
    """Дерево статей отчёта → плоский список (depth, узел) с сохранением вложенности."""
    out = []
    for c in cats:
        out.append((depth, c))
        if c.get("children"):
            out.extend(_flatten_categories(c["children"], depth + 1))
    return out


def pnl_xlsx(report: dict, *, date_from: str, date_to: str) -> bytes:
    periods = report["periods"]
    wb = Workbook()
    ws = wb.active
    ws.title = "ОПиУ"
    method_ru = "метод начисления" if report.get("method") == "accrual" else "кассовый метод"
    ws.cell(row=1, column=1, value=f"Отчёт о прибылях и убытках · {method_ru} · {date_from} — {date_to}").font = BOLD

    hr = 3
    _header_row(ws, hr, ["Статья учёта", *periods, "Итого"])
    r = hr + 1
    last_col = len(periods) + 2

    def section(title: str, sec: dict):
        nonlocal r
        ws.cell(row=r, column=1, value=title).font = BOLD
        for i, p in enumerate(periods):
            _money_cell(ws, r, 2 + i, sec["by_period"][p], bold=True)
        _money_cell(ws, r, last_col, sec["total"], bold=True)
        r += 1
        for depth, c in _flatten_categories(sec["categories"]):
            ws.cell(row=r, column=1, value=("    " * depth) + c["name"])
            for i, p in enumerate(periods):
                _money_cell(ws, r, 2 + i, c["by_period"][p])
            _money_cell(ws, r, last_col, c["total"])
            r += 1

    section("Доходы", report["income"])
    section("Расходы", report["outcome"])

    def total_row(label: str, by_period: dict, total, *, bold=True):
        nonlocal r
        ws.cell(row=r, column=1, value=label).font = BOLD if bold else SUBTLE
        for i, p in enumerate(periods):
            _money_cell(ws, r, 2 + i, by_period[p], bold=bold)
        _money_cell(ws, r, last_col, total, bold=bold)
        r += 1

    total_row("Чистая прибыль", report["profit_by_period"], report["profit_total"])
    total_row("Дивиденды", report["dividends_by_period"], report["dividends_total"], bold=False)
    total_row("Нераспределённая прибыль", report["retained_by_period"], report["retained_total"])

    # Показатели прибыли
    m = report.get("metrics") or {}
    if m:
        r += 1
        ws.cell(row=r, column=1, value="Показатели прибыли").font = BOLD
        r += 1
        for label, key in [
            ("Выручка", "revenue"), ("Прямые расходы", "direct_costs"),
            ("Валовая прибыль", "gross_profit"), ("Косвенные расходы", "indirect_costs"),
            ("Операционная прибыль", "operating_profit"), ("EBITDA", "ebitda"),
            ("Амортизация", "depreciation"), ("EBIT", "ebit"),
            ("Проценты по кредитам", "interest"), ("EBT (прибыль до налога)", "ebt"),
        ]:
            ws.cell(row=r, column=1, value=label)
            _money_cell(ws, r, last_col, m.get(key))
            r += 1

    ws.freeze_panes = "B4"
    _autosize(ws, {1: 42, **{2 + i: 15 for i in range(len(periods))}, last_col: 16})
    return _save(wb)


# ───────────────────────────── ДДС ─────────────────────────────

def cashflow_xlsx(report: dict, *, date_from: str, date_to: str) -> bytes:
    periods = report["periods"]
    wb = Workbook()
    ws = wb.active
    ws.title = "ДДС"
    ws.cell(row=1, column=1, value=f"Движение денежных средств · {date_from} — {date_to}").font = BOLD

    hr = 3
    _header_row(ws, hr, ["Статья учёта", *periods, "Итого"])
    r = hr + 1
    last_col = len(periods) + 2

    def row(label: str, by_period: dict, total=None, *, indent=0, bold=False):
        nonlocal r
        c = ws.cell(row=r, column=1, value=("    " * indent) + label)
        if bold:
            c.font = BOLD
        for i, p in enumerate(periods):
            _money_cell(ws, r, 2 + i, by_period.get(p), bold=bold)
        if total is None:
            total = sum(_num(by_period.get(p)) for p in periods)
        _money_cell(ws, r, last_col, total, bold=bold)
        r += 1

    row("Остаток на начало", report["opening_by_period"], None, bold=True)
    for a in report["activities"]:
        row(a["title"], a["net_by_period"], a["net_total"], bold=True)
        row("Поступления", a["income"]["by_period"], a["income"]["total"], indent=1)
        for c in a["income"]["categories"]:
            row(c["name"], c["by_period"], c["total"], indent=2)
        row("Выплаты", a["outcome"]["by_period"], a["outcome"]["total"], indent=1)
        for c in a["outcome"]["categories"]:
            row(c["name"], c["by_period"], c["total"], indent=2)
    row("Перемещения · списания", report["moves"]["writeoff_by_period"], None, indent=1)
    row("Перемещения · зачисления", report["moves"]["deposit_by_period"], None, indent=1)
    row("Общий денежный поток", report["net_by_period"], report["net_total"], bold=True)
    row("Остаток на конец периода", report["closing_by_period"], None, bold=True)

    ws.freeze_panes = "B4"
    _autosize(ws, {1: 42, **{2 + i: 15 for i in range(len(periods))}, last_col: 16})
    return _save(wb)


# ──────────────────────────── Баланс ────────────────────────────

def balance_xlsx(report: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Баланс"
    ws.cell(row=1, column=1, value=f"Балансовый отчёт на {report['as_of']}").font = BOLD
    if not report.get("balanced"):
        ws.cell(row=1, column=4, value=f"Расхождение: {report.get('difference')}").font = SUBTLE

    _header_row(ws, 3, ["Статья", "Сумма"])
    r = 4

    def block(title: str, total, sections: list[dict]):
        nonlocal r
        ws.cell(row=r, column=1, value=title).font = BOLD
        _money_cell(ws, r, 2, total, bold=True)
        r += 1
        for s in sections:
            ws.cell(row=r, column=1, value="  " + s["title"]).font = BOLD
            _money_cell(ws, r, 2, s["total"], bold=True)
            r += 1
            for it in s["items"]:
                ws.cell(row=r, column=1, value="    " + it["name"])
                _money_cell(ws, r, 2, it["amount"])
                r += 1

    block("АКТИВЫ", report["assets"]["total"], report["assets"]["sections"])
    r += 1
    block("ПАССИВЫ", report["passive_total"],
          report["liabilities"]["sections"] + report["capital"]["sections"])

    ws.freeze_panes = "A4"
    _autosize(ws, {1: 48, 2: 18})
    return _save(wb)


# ─────────────────────────── Операции ───────────────────────────

OPERATIONS_COLUMNS = [
    "Дата оплаты", "Дата начисления", "Тип", "Статус", "Счёт", "Счёт-получатель",
    "Контрагент", "Статья", "Проект", "Сделка", "Сумма", "Валюта", "Комментарий",
]


DEALS_COLUMNS = ["Дата", "Название", "Клиент", "Статус", "Сумма сделки", "Поступило",
                 "Отгружено", "Остаток долга", "Прибыль", "Рентаб., %"]


def deals_xlsx(rows: list[dict], title: str) -> bytes:
    """rows — словари с ключами date/name/client/status/amount/received/shipped/debt/profit/margin."""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    _header_row(ws, 1, DEALS_COLUMNS)
    for ri, r in enumerate(rows, start=2):
        ws.cell(row=ri, column=1, value=r.get("date") or "")
        ws.cell(row=ri, column=2, value=r.get("name") or "")
        ws.cell(row=ri, column=3, value=r.get("client") or "")
        ws.cell(row=ri, column=4, value=r.get("status") or "")
        for col, key in [(5, "amount"), (6, "received"), (7, "shipped"), (8, "debt"), (9, "profit")]:
            _money_cell(ws, ri, col, r.get(key))
        m = r.get("margin")
        ws.cell(row=ri, column=10, value=(f"{m}%" if m is not None else "—")).alignment = RIGHT
    ws.freeze_panes = "A2"
    _autosize(ws, {1: 12, 2: 30, 3: 24, 4: 16, 5: 15, 6: 14, 7: 14, 8: 15, 9: 14, 10: 12})
    return _save(wb)


def operations_xlsx(rows: list[dict]) -> bytes:
    """rows — список словарей с ключами столбцов OPERATIONS_COLUMNS (кроме «Сумма» — число)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Операции"
    _header_row(ws, 1, OPERATIONS_COLUMNS)
    amount_col = OPERATIONS_COLUMNS.index("Сумма") + 1
    for ri, row in enumerate(rows, start=2):
        for ci, key in enumerate(OPERATIONS_COLUMNS, start=1):
            if ci == amount_col:
                _money_cell(ws, ri, ci, row.get(key))
            else:
                ws.cell(row=ri, column=ci, value=row.get(key, ""))
    ws.freeze_panes = "A2"
    _autosize(ws, {1: 12, 2: 14, 3: 13, 4: 10, 5: 18, 6: 18, 7: 22, 8: 22, 9: 16, 10: 16, 11: 16, 12: 8, 13: 30})
    return _save(wb)


# ─────────────────────────── Бюджет (План-Факт) ───────────────────────────

def budget_xlsx(report: dict, *, budget_name: str) -> bytes:
    """Бюджет (БДР/БДДС) в Excel: по статьям × месяцам, План/Факт/Отклонение."""
    periods = report["periods"]
    is_bdr = (report.get("budget_method") or "bdr") == "bdr"
    wb = Workbook()
    ws = wb.active
    ws.title = "БДР" if is_bdr else "БДДС"
    kind_ru = "Бюджет доходов и расходов" if is_bdr else "Бюджет движения денег"
    ws.cell(row=1, column=1, value=f"{kind_ru} · {budget_name}").font = BOLD

    # Заголовок: Статья | <период> план | <период> факт | ... | Итого план | Итого факт | Отклонение
    labels = ["Статья"]
    for p in periods:
        labels += [f"{p} план", f"{p} факт"]
    labels += ["Итого план", "Итого факт", "Отклонение"]
    hr = 3
    _header_row(ws, hr, labels)
    r = hr + 1

    def _row(name, plan_bp, fact_bp, indent=0, bold=False):
        nonlocal r
        c = ws.cell(row=r, column=1, value=("    " * indent) + name)
        if bold:
            c.font = BOLD
        col = 2
        p_tot = f_tot = 0.0
        for p in periods:
            pv, fv = plan_bp.get(p, 0), fact_bp.get(p, 0)
            _money_cell(ws, r, col, pv, bold=bold); _money_cell(ws, r, col + 1, fv, bold=bold)
            p_tot += _num(pv); f_tot += _num(fv); col += 2
        _money_cell(ws, r, col, p_tot, bold=True)
        _money_cell(ws, r, col + 1, f_tot, bold=True)
        _money_cell(ws, r, col + 2, f_tot - p_tot, bold=True)
        r += 1

    for row in report["rows"]:
        _row(row["name"], row["plan_by_period"], row["fact_by_period"])

    ws.freeze_panes = "B4"
    _autosize(ws, {1: 34, **{2 + i: 14 for i in range(len(periods) * 2 + 3)}})
    return _save(wb)


# ─────────────────────────── Платёжный календарь ───────────────────────────

def payment_calendar_xlsx(report: dict) -> bytes:
    """Платёжный календарь в Excel: показатели × периоды."""
    rows = report["rows"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Платёжный календарь"
    ws.cell(row=1, column=1, value="Платёжный календарь").font = BOLD

    labels = [r["period"] for r in rows]
    _header_row(ws, 3, ["Показатель", *labels])
    plan = [
        ("Остаток на начало", "opening"),
        ("Поступления", "income"),
        ("Выплаты", "outcome"),
        ("Денежный поток", "net"),
        ("Остаток на конец", "closing"),
    ]
    r = 4
    for label, key in plan:
        ws.cell(row=r, column=1, value=label).font = BOLD if key in ("net", "closing") else None
        for i, row in enumerate(rows):
            _money_cell(ws, r, 2 + i, row[key], bold=key in ("net", "closing"))
        r += 1
    ws.freeze_panes = "B4"
    _autosize(ws, {1: 22, **{2 + i: 13 for i in range(len(rows))}})
    return _save(wb)
