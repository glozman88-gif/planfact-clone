"""Парсинг загруженных таблиц (xlsx/csv) для мастера импорта операций.

Чистый слой без БД: читает файл в единообразную таблицу строк и предоставляет
парсеры даты/суммы/типа. Построение операций — в app/api/imports.py (нужны БД и модели).
"""
import csv
import io
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

from app.models.enums import OperationType

# Целевые поля операции, на которые пользователь сопоставляет колонки файла.
IMPORT_FIELDS = [
    {"key": "op_date", "label": "Дата оплаты", "required": True},
    {"key": "amount", "label": "Сумма (одной колонкой, со знаком)", "required": False},
    {"key": "amount_income", "label": "Приход (отдельная колонка)", "required": False},
    {"key": "amount_outcome", "label": "Расход (отдельная колонка)", "required": False},
    {"key": "type", "label": "Тип (приход/расход)", "required": False},
    {"key": "account", "label": "Счёт", "required": False},
    {"key": "category", "label": "Статья", "required": False},
    {"key": "counterparty", "label": "Контрагент", "required": False},
    {"key": "project", "label": "Проект", "required": False},
    {"key": "accrual_date", "label": "Дата начисления", "required": False},
    {"key": "description", "label": "Назначение / комментарий", "required": False},
]

DATE_FORMATS = (
    "%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d-%m-%Y",
    "%d.%m.%y", "%Y-%m-%d %H:%M:%S", "%d.%m.%Y %H:%M:%S",
)


def _cell_to_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _decode_text(content: bytes) -> str:
    """Декодирование текстового файла: utf-8 (с BOM) → cp1251 (типично для выписок 1С)."""
    for enc in ("utf-8-sig", "cp1251"):
        try:
            return content.decode(enc)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def parse_1c_exchange(text: str) -> list[list[str]]:
    """Выписка 1С (1CClientBankExchange) → нормализованная таблица.

    Направление определяем по совпадению расчётного счёта выписки (РасчСчет в шапке)
    со счётом плательщика/получателя; запасной вариант — наличие ДатаПоступило/ДатаСписано.
    Колонки результата: Дата, Тип, Сумма, Контрагент, Назначение.
    """
    own_accounts: set[str] = set()
    docs: list[dict[str, str]] = []
    cur: dict[str, str] | None = None
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        key, _, val = line.partition("=")
        key, val = key.strip(), val.strip()
        if key.startswith("СекцияДокумент"):
            cur = {}
        elif key == "КонецДокумента":
            if cur is not None:
                docs.append(cur)
            cur = None
        elif cur is not None:
            cur[key] = val
        elif key in ("РасчСчет", "Расчсчет") and val:
            own_accounts.add(val)

    rows: list[list[str]] = [["Дата", "Тип", "Сумма", "Контрагент", "Назначение"]]
    for d in docs:
        payer = d.get("ПлательщикСчет") or d.get("ПлательщикРасчСчет") or ""
        recv = d.get("ПолучательСчет") or d.get("ПолучательРасчСчет") or ""
        is_income: bool | None = None
        if own_accounts:
            if recv in own_accounts:
                is_income = True
            elif payer in own_accounts:
                is_income = False
        if is_income is None:
            if d.get("ДатаПоступило"):
                is_income = True
            elif d.get("ДатаСписано"):
                is_income = False
        if is_income is None:
            is_income = False
        date = (d.get("ДатаПоступило") if is_income else d.get("ДатаСписано")) or d.get("Дата", "")
        counterparty = (d.get("Плательщик") if is_income else d.get("Получатель")) or ""
        rows.append([
            date,
            "Поступление" if is_income else "Выплата",
            d.get("Сумма", ""),
            counterparty,
            d.get("НазначениеПлатежа", ""),
        ])
    return rows


def read_table(filename: str | None, content: bytes) -> list[list[str]]:
    """Файл → прямоугольная таблица строк (все строки выровнены по ширине).

    Поддержка: .xlsx/.xlsm (openpyxl), выписка 1С (1CClientBankExchange), CSV.
    """
    name = (filename or "").lower()
    rows: list[list[str]] = []
    if name.endswith((".xlsx", ".xlsm")):
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            rows.append([_cell_to_str(v) for v in row])
        wb.close()
    else:
        raw = _decode_text(content)
        if raw.lstrip().startswith("1CClientBankExchange"):
            rows = parse_1c_exchange(raw)
        else:
            try:
                dialect = csv.Sniffer().sniff(raw[:2048], delimiters=";,\t") if raw.strip() else csv.excel
            except csv.Error:
                dialect = csv.excel
            rows = [[(c or "").strip() for c in r] for r in csv.reader(io.StringIO(raw), dialect=dialect)]
    # отбрасываем полностью пустые строки и выравниваем ширину
    rows = [r for r in rows if any(c for c in r)]
    width = max((len(r) for r in rows), default=0)
    return [r + [""] * (width - len(r)) for r in rows]


def parse_date(s: str) -> date | None:
    s = (s or "").strip()
    if not s:
        return None
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def parse_amount(s: str) -> Decimal | None:
    s = (s or "").strip().replace("\xa0", "").replace(" ", "")
    if not s:
        return None
    neg = s.startswith("-") or s.startswith("(") and s.endswith(")")
    s = s.replace(",", ".")
    s = re.sub(r"[^0-9.]", "", s)
    if not s or s == ".":
        return None
    try:
        val = Decimal(s)
    except (InvalidOperation, ValueError):
        return None
    return -val if neg else val


def parse_type(s: str) -> OperationType | None:
    """Тип операции из текстовой колонки. None — не распознан (решит знак суммы)."""
    s = (s or "").strip().lower()
    if not s:
        return None
    if s.startswith(("in", "доход", "прих", "поступ", "+", "кредит", "credit")):
        return OperationType.income
    if s.startswith(("out", "расход", "выпл", "списан", "-", "дебет", "debit", "exp")):
        return OperationType.outcome
    return None
